import openpyxl, collections, urllib, sys, re, pymongo

filename = openpyxl.load_workbook('C:\Users\karang\Desktop\costing.xlsx')

colo = filename.get_sheet_by_name('18390')
ggva = filename.get_sheet_by_name('24946')


mongoserver = "app82.glam.colo"
mongoport = 27017
mydb = "sojourner"
mycollection = "sojourner"

data = collections.defaultdict(list)

index = 1
price = list()
newindex = 2

for row in colo.iter_rows():  
    if newindex < 1696:
        if colo['L' + str(newindex)].value is None:
            if colo['D' + str(newindex)].value:
                if re.findall('RAM', colo['D' + str(newindex)].value):
                    price.append(colo['F' + str(newindex)].value)    
        if colo['L' + str(newindex)].value is not None:
            data[index] = list()
            mysum = 0
            for p in price:
                if p == None:
                    p = 0
                mysum = mysum + p
            data[index-1].append(mysum)    
            price = list()     
            data[index].append(colo['L' + str(newindex)].value)
            data[index].append((colo['J' + str(newindex)].value).strftime('%m/%d/%Y'))
            data[index].append((colo['K' + str(newindex)].value).strftime('%m/%d/%Y'))
            price.append(colo['F' + str(newindex)].value)    
            index = index + 1
        newindex = newindex + 1

data[index-1].append(price)

previoushost = ""
newindex = 2

for row in ggva.iter_rows():
    if newindex < 697:
        if ggva['J' + str(newindex)].value is not None:
            if ggva['D' + str(newindex)].value != 0:
                if ggva['J' + str(newindex)].value == previoushost:
                    data[index-1].append(ggva['D' + str(newindex)].value)
                else:
                    data[index] = list()
                    data[index].append(ggva['J' + str(newindex)].value)
                    if ggva['H' + str(newindex)].value is not None:
                        data[index].append((ggva['H' + str(newindex)].value).strftime('%m/%d/%Y'))
                    else:
                        data[index].append(ggva['H' + str(newindex)].value)
                    if ggva['I' + str(newindex)].value is not None:
                        data[index].append((ggva['I' + str(newindex)].value).strftime('%m/%d/%Y'))
                    else:
                        data[index].append(ggva['I' + str(newindex)].value)
                    data[index].append(ggva['D' + str(newindex)].value)
                index = index + 1
            previoushost = ggva['J' + str(newindex)].value
        newindex = newindex + 1

newdata = collections.defaultdict(dict)
dccosting = list()

for key, val in data.iteritems():
    if len(val) > 1:
        if not newdata.has_key(val[0]):
            if val[0] != 'SPaccess':
                if val[1]:
                    newdata[val[0]]['commission_date'] = val[1]
                else:
                    newdata[val[0]]['commission_date'] = ''
                if val[2]:
                    newdata[val[0]]['termination_date'] = val[2]
                else:
                    newdata[val[0]]['termination_date'] = ''
                if val[3]:
                    mysum = 0
                    if type(val[3]) is list:
                        for c in val[3]:
                            if c == None:
                                c = 0
                            mysum = c + mysum
                        dccosting.append(float(mysum))
                        newdata[val[0]]['cost'] = '$' + str(float(mysum))
                    else:
                        dccosting.append(float(val[3]))
                        newdata[val[0]]['cost'] = '$' + str(float(val[3]))
                else:
                    newdata[val[0]]['cost'] = ''
    
try:
    conn = pymongo.MongoClient(mongoserver, mongoport)
except pymongo.errors.ConnectionFailure, e:
    print "Could not connect to MongoDB: %s" % e

db = conn[mydb]
collection = db[mycollection]

for key, val in newdata.iteritems():
    result = collection.find_one({'_id': key})
    if result is not None:
        for k, v in val.iteritems():
            collection.update({'_id': key}, {'$set': {k: v}})
